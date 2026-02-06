import openpyxl
from openpyxl.styles import PatternFill
import os
import glob

# --- CONFIGURATION ---
SECTION_FOLDER_PATH = r"C:\Users\Jemar John\Downloads\Sectioning Lists"

# Columns to Clear Content: J(10) to O(15)
COLS_TO_CLEAN = range(10, 16) # 10, 11, 12, 13, 14, 15

def reset_and_clean_files():
    print(f"Resetting files in: {SECTION_FOLDER_PATH}")
    
    files = glob.glob(os.path.join(SECTION_FOLDER_PATH, "*.xlsx"))
    
    if not files:
        print("No Excel files found!")
        return

    # Define "No Fill" style
    no_fill = PatternFill(fill_type=None)

    for file_path in files:
        if "~$" in file_path: continue # Skip temp files
        
        file_name = os.path.basename(file_path)
        print(f"Processing: {file_name}...", end="")
        
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # 1. Reset Sheet Format Defaults
            ws.sheet_format.defaultRowHeight = 15
            ws.sheet_format.defaultColWidth = 8.43

            # 2. Reset Row Dimensions (Height)
            # We create a list of keys first to avoid error while modifying dictionary during iteration
            for row_idx in list(ws.row_dimensions.keys()):
                del ws.row_dimensions[row_idx] # Deleting entry forces default height

            # 3. Reset Column Dimensions (Width)
            for col_idx in list(ws.column_dimensions.keys()):
                del ws.column_dimensions[col_idx] # Deleting entry forces default width

            # 4. Iterate ALL cells to Remove Color and Clear J-O
            for row in ws.iter_rows():
                for cell in row:
                    # Remove Background Color
                    cell.fill = no_fill
                    
                    # Clear Content if in Columns J-O
                    # cell.column is 1-based index (A=1, J=10)
                    if cell.column in COLS_TO_CLEAN:
                        cell.value = None

            wb.save(file_path)
            print(" Done (Cleaned & Reset).")
            
        except Exception as e:
            print(f"\n  Error processing {file_name}: {e}")

    print("\nAll files have been reset to default formatting and cleaned.")

if __name__ == "__main__":
    reset_and_clean_files()