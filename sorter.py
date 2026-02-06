import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import re
import unicodedata
import os
import glob

# --- CONFIGURATION ---
AER_FILE_PATH = r"C:\Users\Jemar John\Downloads\MSU-IIT AER 2S25-26 DATA.xlsx"
SECTION_FOLDER_PATH = r"C:\Users\Jemar John\Downloads\Sectioning Lists"
AER_OUTPUT_PATH = r"C:\Users\Jemar John\Downloads\MSU-IIT AER_Highlighted.xlsx"

# Sheet Name in AER (set to None to auto-detect active sheet)
AER_SHEET_NAME = "ROSTER"

# Colors (Hex codes)
COLOR_BLUE = "8CB5F9"  # Found in AER
COLOR_RED = "FF9999"   # Not found in AER (using a lighter red for readability)

def normalize_name(text):
    """
    Cleans a string: lowercase, removes special chars, splits into tokens.
    Example: "Catian, Freddie Jr." -> {'catian', 'freddie', 'jr'}
    """
    if not isinstance(text, str) or not text:
        return set()
    
    # Normalize unicode (e.g., ñ -> n)
    text = unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('utf-8')
    # Keep only letters and numbers
    text = re.sub(r'[^a-z0-9\s]', ' ', text.lower())
    # Return a set of words
    return set(text.split())

def is_match(tokens1, tokens2):
    """
    Checks if two sets of name tokens match.
    Logic: One set is a subset of the other, AND they share at least 2 words.
    """
    if not tokens1 or not tokens2:
        return False
    common = tokens1.intersection(tokens2)
    # Must share at least 2 parts (e.g. First + Last) to count as a match
    return len(common) >= 2 and (tokens1.issubset(tokens2) or tokens2.issubset(tokens1))

def get_aer_names(file_path, sheet_name):
    """Reads the AER file and returns a list of name token sets."""
    print(f"Reading AER Master List: {file_path}")
    try:
        # Read using pandas for speed first
        df = pd.read_excel(file_path, sheet_name=sheet_name if sheet_name else 0)
        
        aer_tokens_list = []
        # Assuming AER structure: Col B (Last Name), Col C (First Name)
        # We iterate assuming standard DataFrame logic or specific column names if known.
        # Based on previous files, rows 12+ contain data. We'll iterate strictly.
        
        # Simpler approach: Iterate generic rows and look for name-like structures
        # or use specific columns if the format is strict.
        # Let's rely on column indices 1 (Last) and 2 (First) like the previous script.
        
        # We'll use openpyxl to read to be safe about skipped rows
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if not row or len(row) < 3: continue
            lname, fname = row[1], row[2] # Col B and C
            
            # Simple check if this is a header row
            if str(lname).lower() == "l-name": continue
            
            if lname:
                full_name = f"{lname} {fname if fname else ''}"
                tokens = normalize_name(full_name)
                if tokens:
                    aer_tokens_list.append(tokens)
                    
        print(f"Loaded {len(aer_tokens_list)} names from AER.")
        return aer_tokens_list
        
    except Exception as e:
        print(f"Error reading AER names: {e}")
        return []

def process_section_files(folder_path, aer_tokens_list):
    """
    Iterates through all Excel files in the folder.
    Highlights RED if not in AER.
    Returns a collected list of ALL students found in sections (for the blue highlight step).
    """
    print(f"\nScanning Section Files in: {folder_path}")
    all_section_tokens = []
    
    # Get all xlsx files
    files = glob.glob(os.path.join(folder_path, "*.xlsx"))
    
    if not files:
        print("No Excel files found in the sectioning folder!")
        return []

    red_fill = PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type="solid")

    for file_path in files:
        if "~$" in file_path: continue # Skip temp files
        
        print(f"Processing: {os.path.basename(file_path)}...")
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            modified = False
            
            # Find Name Column (Look for 'studentname' header)
            name_col_idx = 3 # Default to column D (index 3)
            header_row_idx = 4 # Default to row 5 (index 4)
            
            # Smart header detection
            for r_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
                for c_idx, cell_val in enumerate(row):
                    if str(cell_val).lower().replace(" ","") in ['studentname', 'name', 'fullname']:
                        name_col_idx = c_idx
                        header_row_idx = r_idx + 1 # 1-based for openpyxl
                        break
            
            # Iterate rows strictly after header
            for row in ws.iter_rows(min_row=header_row_idx + 1):
                if not row or len(row) <= name_col_idx: continue
                
                cell_name = row[name_col_idx]
                if not cell_name.value: continue
                
                sec_tokens = normalize_name(str(cell_name.value))
                if not sec_tokens: continue
                
                # Check against AER
                found_in_aer = False
                for aer_tokens in aer_tokens_list:
                    if is_match(aer_tokens, sec_tokens):
                        found_in_aer = True
                        break
                
                # ACTION: If NOT in AER -> Highlight RED
                if not found_in_aer:
                    for cell in row: # Highlight the whole row
                        cell.fill = red_fill
                    modified = True
                else:
                    # Collect confirmed student for the Blue Step
                    all_section_tokens.append(sec_tokens)
            
            if modified:
                wb.save(file_path) # Overwrite the section file
                print(f"  -> Saved changes (Red highlights applied).")
                
        except Exception as e:
            print(f"  -> Error processing file: {e}")
            
    return all_section_tokens

def highlight_aer_master(aer_path, output_path, section_tokens_list, sheet_name):
    """Highlights AER students in BLUE if they exist in the collected section list."""
    print(f"\nFinalizing Master List...")
    
    try:
        wb = openpyxl.load_workbook(aer_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        
        blue_fill = PatternFill(start_color=COLOR_BLUE, end_color=COLOR_BLUE, fill_type="solid")
        count = 0
        
        for row in ws.iter_rows(min_row=1):
            cell_lname = row[1] # Col B
            cell_fname = row[2] # Col C
            
            if not cell_lname.value: continue
            
            # Get tokens for this AER student
            full_name = f"{cell_lname.value} {cell_fname.value if cell_fname.value else ''}"
            aer_tokens = normalize_name(full_name)
            
            if "l-name" in aer_tokens: continue # Skip header
            
            # Check if this student appeared in ANY section file
            found_in_section = False
            for sec_tokens in section_tokens_list:
                if is_match(aer_tokens, sec_tokens):
                    found_in_section = True
                    break
            
            if found_in_section:
                cell_lname.fill = blue_fill
                count += 1
                
        wb.save(output_path)
        print(f"Success! {count} students highlighted BLUE in the Master List.")
        print(f"Master file saved to: {output_path}")
        
    except Exception as e:
        print(f"Error highlighting master list: {e}")

def main():
    # 1. Load AER Names
    aer_names = get_aer_names(AER_FILE_PATH, AER_SHEET_NAME)
    
    if not aer_names:
        print("CRITICAL: No names found in AER file. Stopping.")
        return

    # 2. Process Section Files (Highlight Red & Collect Names)
    found_section_students = process_section_files(SECTION_FOLDER_PATH, aer_names)
    
    # 3. Highlight AER File (Blue)
    highlight_aer_master(AER_FILE_PATH, AER_OUTPUT_PATH, found_section_students, AER_SHEET_NAME)
    
    print("\n--- JOB COMPLETE ---")

if __name__ == "__main__":
    main()